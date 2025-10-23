from io import BytesIO
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# --- Styling Constants ---
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
TITLE_FONT = Font(size=18, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
CURRENCY_FORMAT = '$#,##0.00'

def _add_startup_activities_sheet(wb, activities):
    """Adds the Startup Activities sheet to the workbook."""
    ws = wb.create_sheet(title="Startup Activities")
    ws['A1'] = 'Startup Activities'
    ws['A1'].font = TITLE_FONT
    ws['A1'].fill = TITLE_FILL
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

    headers = ['Activity', 'Description', 'Weight (%)', 'Progress (%)']
    ws.append(headers)
    for cell in ws[2]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for activity in activities:
        ws.append([
            activity.get('activity'),
            activity.get('description'),
            activity.get('weight'),
            activity.get('progress')
        ])

def _add_revenue_sheet(wb, products, seasonality_factors, company_name):
    """Adds the Quarterly Revenue sheet and chart to the workbook."""
    ws = wb.create_sheet(title="Quarterly Revenue", index=0)
    
    # Title
    display_company_name = company_name if company_name else 'My Awesome Startup'
    ws['A1'] = f'Financial Forecast for {display_company_name}'
    ws['A1'].font = TITLE_FONT
    ws['A1'].fill = TITLE_FILL

    # Headers
    product_names = [p.get('description', 'N/A') for p in products]
    headers = ['Quarter'] + product_names + ['Total Revenue']
    ws.append(headers)
    for cell in ws[2]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    # --- Data Calculation ---
    total_factor = sum(seasonality_factors)
    normalized_factors = [(f / total_factor) * 12 for f in seasonality_factors] if total_factor > 0 else [1.0] * 12

    product_monthly_revenues = []
    for p in products:
        price = float(p.get('price', 0) or 0)
        volume = int(p.get('sales_volume', 0) or 0)
        unit = p.get('sales_volume_unit', 'monthly')
        annual_volume = volume * 12 if unit == 'monthly' else volume * 4
        product_monthly_revenues.append((price * annual_volume) / 12)

    # Populate quarterly data
    for q in range(4):
        row_data = [f'Q{q+1}']
        quarterly_total = 0
        for monthly_rev in product_monthly_revenues:
            quarterly_prod_rev = sum(monthly_rev * normalized_factors[q * 3 + i] for i in range(3))
            row_data.append(quarterly_prod_rev)
            quarterly_total += quarterly_prod_rev
        row_data.append(quarterly_total)
        ws.append(row_data)

    # Formatting
    for row in ws.iter_rows(min_row=3, min_col=2, max_col=ws.max_column):
        for cell in row:
            cell.number_format = CURRENCY_FORMAT

    # --- Chart ---
    chart = BarChart()
    chart.title = "Quarterly Revenue"
    chart.y_axis.title = "Revenue"
    chart.x_axis.title = "Quarter"
    chart.y_axis.number_format = CURRENCY_FORMAT
    chart.grouping = "clustered" # Changed from "stacked"

    # Include all revenue columns, including the total
    data = Reference(ws, min_col=2, min_row=2, max_col=ws.max_column, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=3, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "A8")
    return product_monthly_revenues

def _add_pnl_sheet(wb, product_monthly_revenues, operating_expenses, cogs_percentage, loan_details, depreciation, interest_expense):
    """Adds the 5-Year P&L Summary sheet and chart."""
    ws = wb.create_sheet(title="Annual P&L Summary")
    ws['A1'] = 'Profit & Loss Summary (USD)'
    ws['A1'].font = TITLE_FONT
    ws['A1'].fill = TITLE_FILL
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)

    headers = ['Year', 'Total Revenue', 'COGS', 'Gross Profit', 'Operating Expenses', 'Net Operating Income', 'Depreciation', 'Earnings Before Tax', 'Taxes', 'Net Income', 'DSCR']
    ws.append(headers)
    for cell in ws[2]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    # --- P&L Calculation ---
    y1_revenue = sum(product_monthly_revenues) * 12
    y1_opex = sum((float(e.get('amount', 0)) * 12 if e.get('frequency') == 'monthly' else float(e.get('amount', 0)) * 4) for e in operating_expenses)

    current_revenue, current_opex = y1_revenue, y1_opex
    for year in range(1, 6):
        if year > 1:
            current_revenue *= 1.10  # 10% revenue growth
            current_opex *= 1.05     # 5% opex growth

        cogs = current_revenue * (cogs_percentage / 100)
        gross_profit = current_revenue - cogs
        noi = gross_profit - current_opex
        ebt = noi - depreciation - interest_expense
        taxes = max(0, ebt * 0.25) # 25% standard tax assumption
        net_income = ebt - taxes
        
        total_debt_service = (loan_details.get('monthly_payment', 0) or 0) * 12
        dscr = (noi / total_debt_service) if total_debt_service > 0 else 0

        ws.append([year, current_revenue, cogs, gross_profit, current_opex, noi, depreciation, ebt, taxes, net_income, dscr if dscr > 0 else 'N/A'])

    # Formatting
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=2, max_col=10):
        for cell in row:
            cell.number_format = CURRENCY_FORMAT
    for cell in ws['K']:
        if cell.row > 2: cell.number_format = '0.00'

    # --- Chart ---
    chart = BarChart()
    chart.style = 13
    chart.grouping = "stacked"
    chart.title = "5-Year Financial Projections"
    chart.y_axis.title = "Amount (USD)"
    chart.x_axis.title = "Year"
    chart.y_axis.number_format = CURRENCY_FORMAT

    cats = Reference(ws, min_col=1, min_row=3, max_row=ws.max_row)
    chart.set_categories(cats)

    # Add data series for 'Total Revenue', 'Gross Profit', and 'Net Income'
    data_cols = [2, 4, 10] 
    for col in data_cols:
        data = Reference(ws, min_col=col, min_row=2, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)

    ws.add_chart(chart, "A10")

def _add_loan_sheet(wb, loan_details):
    """Adds the Loan Payment Schedule sheet and chart if data is available."""
    if not loan_details or not loan_details.get('schedule'):
        return

    ws = wb.create_sheet(title="Loan Payment Schedule")
    ws['A1'] = 'Loan Payment Schedule'
    ws['A1'].font = TITLE_FONT
    ws['A1'].fill = TITLE_FILL
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)

    # Summary
    ws.append(['Loan Amount', loan_details.get('loan_amount')])
    ws.append(['Annual Interest Rate (%)', loan_details.get('interest_rate')])
    ws.append(['Loan Term (Years)', loan_details.get('loan_term')])
    ws.append(['Monthly Payment', loan_details.get('monthly_payment')])
    for row_num in range(3, 7):
        ws[f'B{row_num}'].number_format = CURRENCY_FORMAT if row_num in [3, 6] else '0.00'

    # Schedule Table
    ws.append([]) # Spacer
    headers = ['Month', 'Principal', 'Interest', 'Remaining Balance']
    ws.append(headers)
    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for item in loan_details['schedule']:
        ws.append([item['month'], item['principal_payment'], item['interest_payment'], item['remaining_balance']])
    
    for row in ws.iter_rows(min_row=header_row + 1, min_col=2, max_col=4):
        for cell in row:
            cell.number_format = CURRENCY_FORMAT

    # --- Chart ---
    #chart = BarChart()
    #chart.title = "Loan Payments (Principal vs. Interest)"
    #chart.y_axis.title = "Payment Amount"
    #chart.y_axis.number_format = CURRENCY_FORMAT
    #chart.grouping = "stacked"

    #data = Reference(ws, min_col=2, min_row=header_row, max_col=3, max_row=ws.max_row)
    #cats = Reference(ws, min_col=1, min_row=header_row + 1, max_row=ws.max_row)
    #chart.add_data(data, titles_from_data=True)
    #chart.set_categories(cats)
    #ws.add_chart(chart, "F2")

def _finalize_workbook(wb):
    """Applies final formatting to all sheets in the workbook."""
    for sheet in wb.worksheets:
        # Adjust column widths
        for col_idx in range(1, sheet.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for cell in sheet[col_letter]:
                try:
                    val_len = len(str(cell.value))
                    if val_len > max_len:
                        max_len = val_len
                except: pass
            sheet.column_dimensions[col_letter].width = max_len + 2
        
        # Special case for merged title cells
        if sheet.title in ["Quarterly Revenue", "Annual P&L Summary", "Startup Activities"]:
            sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=sheet.max_column)

def create_forecast_spreadsheet(products, operating_expenses, cogs_percentage, loan_details, seasonality_factors, company_name, depreciation, interest_expense, startup_activities):
    """Creates an Excel spreadsheet with financial forecast and loan amortization data."""
    if seasonality_factors is None:
        seasonality_factors = [1.0] * 12

    wb = Workbook()
    wb.remove(wb.active) # Remove default sheet

    # Add sheets
    product_monthly_revenues = _add_revenue_sheet(wb, products, seasonality_factors, company_name)
    _add_pnl_sheet(wb, product_monthly_revenues, operating_expenses, cogs_percentage, loan_details, depreciation, interest_expense)
    _add_loan_sheet(wb, loan_details)
    _add_startup_activities_sheet(wb, startup_activities)

    # Final formatting
    _finalize_workbook(wb)

    # Save to an in-memory file
    in_memory_file = BytesIO()
    wb.save(in_memory_file)
    in_memory_file.seek(0)
    return in_memory_file